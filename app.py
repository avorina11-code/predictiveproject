# -*- coding: utf-8 -*-
"""
=============================================================================
 CONTROL ROOM WFM — PILOTAGE PRODUCTION TEMPS RÉEL & PRÉDICTIF
=============================================================================
Application Streamlit (fichier unique) combinant :
  - un moteur métier Python (Pandas / NumPy) : cumuls, projections WFM,
    détection du "Point de Non-Retour" sur le Service Level, diagnostic
    automatique de cause racine, actions prescriptives ;
  - des composants d'interface HTML5 / CSS3 / JavaScript injectés via
    `streamlit.components.v1.html` pour un rendu "Control Room" (bandeau
    d'alerte animé néon, jauge Chart.js, panneau "Vigie IA" à effet
    machine à écrire, courbe de projection interactive) ;
  - un tableau de bord Pandas avec mise en forme conditionnelle (heatmap).

Auteur   : Expert Senior WFM / Data Science / Full-Stack (Python + JS)
Fichier  : app.py (unique, autonome)
=============================================================================
"""

import re
import io
import json
import string
import datetime as dt

import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. CONFIGURATION GÉNÉRALE DE LA PAGE STREAMLIT
# =============================================================================
st.set_page_config(
    page_title="Control Room WFM — Pilotage Temps Réel",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- Thème visuel global "Control Room" (fond sombre, accents néon) --------
st.markdown(
    """
    <style>
    .stApp { background-color: #070b12; }
    section[data-testid="stSidebar"] { background-color: #0c111c; }
    h1, h2, h3, h4 { color: #e6f1ff !important; font-family: 'Segoe UI', sans-serif; }
    .stMetric { background-color: #0f1626; border-radius: 12px; padding: 10px; }
    div[data-testid="stMetricValue"] { color: #00e5ff; }
    hr { border-color: #1c2536; }
    .bloc-titre {
        color:#8fa3c7; font-size:13px; letter-spacing:2px; text-transform:uppercase;
        font-weight:700; margin-top:18px; margin-bottom:6px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# =============================================================================
# 2. UTILITAIRES DE NORMALISATION DES DONNÉES
# =============================================================================
def normaliser_nom_colonne(col: str) -> str:
    """Convertit un nom de colonne brut (avec accents/espaces/%) en un
    identifiant technique stable, indépendant de l'encodage du fichier."""
    col = str(col).strip().upper()
    remplacements = {
        "É": "E", "È": "E", "Ê": "E", "Ë": "E",
        "À": "A", "Â": "A", "Ô": "O", "Î": "I", "Ï": "I", "Ù": "U", "Û": "U", "Ç": "C",
    }
    for a, b in remplacements.items():
        col = col.replace(a, b)
    col = col.replace("%", "PCT")
    col = re.sub(r"[^A-Z0-9]+", "_", col)
    col = re.sub(r"_+", "_", col).strip("_")
    return col


# Mapping "nom normalisé attendu" -> "nom interne utilisé dans le code"
MAP_COLONNES_PROD = {
    "TRANCHE": "tranche",
    "RECUS": "recus",
    "TRAITES": "traites",
    "PREVISION": "prevision",
    "TRP_PCT": "trp",
    "QS_PCT": "qs",
    "SL_PCT": "sl",
    "DMC_S": "dmc",
    "ACW_S": "acw",
    "DMT_S": "dmt",
    "CONNECTES": "connectes",
    "EN_TRAIT": "en_traitement",
    "DISPO_PCT": "dispo",
    "ABAND_MOY": "aband_moy",
    "BESOIN": "besoin",
    "PLANNING": "planning",
    "ECART": "ecart",
}

MAP_COLONNES_PLANNING = {
    "LOGIN_VOCALCOM": "login",
    "CODE_RH": "code_rh",
    "NOM_PRENOM": "nom_prenom",
    "HEURE_DEBUT": "heure_debut",
    "HEURE_FIN": "heure_fin",
    "PAUSE_DEBUT": "pause_debut",
    "PAUSE_FIN": "pause_fin",
}


def parse_heure(val):
    """Parse une valeur hétérogène (str 'HH:MM', datetime.time, datetime.datetime,
    NaN) et retourne un objet datetime.time ou None."""
    if val is None:
        return None
    if isinstance(val, float) and np.isnan(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.time()
    if isinstance(val, dt.datetime):
        return val.time()
    if isinstance(val, dt.time):
        return val
    s = str(val).strip()
    if s == "" or s.lower() in ("nan", "none", "nat"):
        return None
    m = re.search(r"(\d{1,2})[:hH](\d{2})", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return dt.time(h, mi)
    return None


def normaliser_pourcentage(serie: pd.Series) -> pd.Series:
    """Ramène une colonne de pourcentage sur une échelle 0-100, qu'elle soit
    saisie en fraction (0.85) ou déjà en pourcentage (85)."""
    serie = pd.to_numeric(serie, errors="coerce")
    if serie.dropna().empty:
        return serie.fillna(0.0)
    if serie.dropna().max() <= 1.5:
        serie = serie * 100.0
    return serie.fillna(0.0)


def charger_fichier(uploaded_file, map_colonnes: dict) -> pd.DataFrame:
    """Charge un fichier Excel/CSV uploadé et renomme ses colonnes selon le
    mapping fourni. Les colonnes non reconnues sont conservées telles quelles."""
    if uploaded_file.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded_file, sep=None, engine="python")
    else:
        df = pd.read_excel(uploaded_file)
    df.columns = [normaliser_nom_colonne(c) for c in df.columns]
    df = df.rename(columns={k: v for k, v in map_colonnes.items() if k in df.columns})
    return df


# =============================================================================
# 2 bis. MODÈLES DE FICHIERS TÉLÉCHARGEABLES (templates Excel)
# =============================================================================
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_EXEMPLE_FONT = Font(name="Arial", italic=True, color="6B7280", size=10)
_LEGENDE_TITRE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_LEGENDE_TITRE_FILL = PatternFill("solid", fgColor="374151")
_BORDURE = Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def _entete_feuille(ws, colonnes: list):
    """Écrit une ligne d'en-tête stylée (fond sombre, texte blanc, bordures)."""
    for j, col in enumerate(colonnes, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDURE
        ws.column_dimensions[get_column_letter(j)].width = max(13, len(col) + 4)
    ws.row_dimensions[1].height = 28


def _ligne_exemple(ws, valeurs: list, ligne: int = 2):
    """Écrit une ligne d'exemple en italique grisé pour montrer le format attendu."""
    for j, val in enumerate(valeurs, start=1):
        cell = ws.cell(row=ligne, column=j, value=val)
        cell.font = _EXEMPLE_FONT
        cell.border = _BORDURE


def _feuille_legende(wb, lignes: list, titre_feuille: str = "Légende"):
    """Ajoute une feuille 'Légende' décrivant chaque colonne du modèle."""
    ws = wb.create_sheet(titre_feuille)
    entetes = ["Colonne", "Description"]
    for j, e in enumerate(entetes, start=1):
        cell = ws.cell(row=1, column=j, value=e)
        cell.font = _LEGENDE_TITRE_FONT
        cell.fill = _LEGENDE_TITRE_FILL
        cell.border = _BORDURE
    for i, (col, desc) in enumerate(lignes, start=2):
        c1 = ws.cell(row=i, column=1, value=col)
        c2 = ws.cell(row=i, column=2, value=desc)
        c1.font = Font(name="Arial", bold=True, size=10)
        c2.font = Font(name="Arial", size=10)
        c1.alignment = Alignment(vertical="top")
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        c1.border = _BORDURE
        c2.border = _BORDURE
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 85
    ws.row_dimensions[1].height = 22


def generer_modele_production() -> bytes:
    """Génère (en mémoire) le classeur modèle du fichier 1 — Métriques de
    production réelles par tranche — avec en-tête stylé, une ligne d'exemple
    et un onglet Légende expliquant chaque colonne."""
    colonnes = [
        "TRANCHE", "REÇUS", "TRAITÉS", "PRÉVISION", "TRP %", "QS %", "SL %",
        "DMC (S)", "ACW (S)", "DMT (S)", "CONNECTÉS", "EN TRAIT.", "DISPO %",
        "ABAND. MOY", "BESOIN", "PLANNING", "ÉCART",
    ]
    exemple = ["08:00", 62, 60, 58, 96.5, 92.0, 88.0, 250, 14, 265, 20, 18, 78.0, 3.2, 21, 21, -1]

    wb = Workbook()
    ws = wb.active
    ws.title = "Production"
    _entete_feuille(ws, colonnes)
    _ligne_exemple(ws, exemple)
    ws.freeze_panes = "A2"

    legende = [
        ("TRANCHE", "Heure de début de la tranche de 30 min, format HH:MM (ex : 08:00)."),
        ("REÇUS", "Nombre d'appels/contacts reçus sur la tranche."),
        ("TRAITÉS", "Nombre d'appels/contacts traités sur la tranche."),
        ("PRÉVISION", "Volume prévisionnel d'appels pour la tranche (issu du forecast WFM)."),
        ("TRP %", "Taux de Réponse (%) sur la tranche — saisir en valeur (ex : 96.5) ou en fraction (0.965)."),
        ("QS %", "Qualité de Service (%) sur la tranche."),
        ("SL %", "Service Level (%) réalisé sur la tranche."),
        ("DMC (S)", "Durée Moyenne de Conversation, en secondes."),
        ("ACW (S)", "After Call Work — post-appel, en secondes."),
        ("DMT (S)", "Durée Moyenne de Traitement (DMC + ACW), en secondes."),
        ("CONNECTÉS", "Nombre d'agents effectivement connectés sur la tranche."),
        ("EN TRAIT.", "Nombre d'agents en traitement d'appel au même instant."),
        ("DISPO %", "Taux de disponibilité des agents (%)."),
        ("ABAND. MOY", "Durée moyenne avant abandon des appelants, en secondes."),
        ("BESOIN", "Effectif requis théorique (Erlang) pour la tranche."),
        ("PLANNING", "Effectif planifié (issu du planning) sur la tranche."),
        ("ÉCART", "Écart Connectés − Planning."),
    ]
    _feuille_legende(wb, legende)

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def generer_modele_planning() -> bytes:
    """Génère (en mémoire) le classeur modèle du fichier 2 — Planning agents
    & pauses déjeuner — avec en-tête stylé, deux lignes d'exemple et un
    onglet Légende."""
    colonnes = ["Login_Vocalcom", "Code_RH", "Nom_Prenom", "Heure_Debut", "Heure_Fin", "Pause_Debut", "Pause_Fin"]
    exemples = [
        ["AG001", "RH1000", "Dupont Marie", "08:00", "17:00", "12:00", "12:45"],
        ["AG002", "RH1001", "Martin Julien", "07:30", "16:30", "11:30", "12:15"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"
    _entete_feuille(ws, colonnes)
    for i, ex in enumerate(exemples, start=2):
        _ligne_exemple(ws, ex, ligne=i)
    ws.freeze_panes = "A2"

    legende = [
        ("Login_Vocalcom", "Identifiant de connexion de l'agent dans l'outil de téléphonie (Vocalcom)."),
        ("Code_RH", "Matricule / code RH de l'agent."),
        ("Nom_Prenom", "Nom et prénom de l'agent."),
        ("Heure_Debut", "Heure de prise de poste, format HH:MM."),
        ("Heure_Fin", "Heure de fin de poste, format HH:MM."),
        ("Pause_Debut", "Heure de début de la pause déjeuner, format HH:MM (laisser vide si aucune pause)."),
        ("Pause_Fin", "Heure de fin de la pause déjeuner, format HH:MM (laisser vide si aucune pause)."),
    ]
    _feuille_legende(wb, legende)

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


# =============================================================================
# 3. DONNÉES DE DÉMONSTRATION (utilisées si aucun fichier n'est chargé)
# =============================================================================
def generer_demo_production() -> pd.DataFrame:
    """Génère un jeu de données réaliste sur une journée (08:00-18:00, pas de
    30 min) avec un scénario de dérive DMT provoquant un point de non-retour
    en milieu de journée — utile pour démontrer le fonctionnement de l'outil."""
    heures = pd.date_range("08:00", "17:30", freq="30min").time
    rng = np.random.default_rng(42)
    lignes = []
    dmt_cible_demo = 280
    for i, h in enumerate(heures):
        prevision = int(60 + 25 * np.sin(i / 3) + rng.integers(-5, 5))
        prevision = max(prevision, 20)
        # Dérive DMT progressive à partir de la tranche 6 (11:00)
        derive_dmt = 1.0 if i < 6 else 1.0 + min(0.35, (i - 5) * 0.05)
        dmt = dmt_cible_demo * derive_dmt + rng.integers(-8, 8)
        recus = int(prevision * rng.uniform(0.92, 1.12))
        planning = int(prevision / 3.2) + rng.integers(-1, 2)
        connectes = max(planning - (2 if i >= 6 else 0) - rng.integers(0, 2), 1)
        traites = int(recus * rng.uniform(0.9, 1.0))
        sl = max(30, min(98, 95 - (dmt - dmt_cible_demo) * 0.22 + rng.integers(-3, 3)))
        qs = max(50, min(99, sl + rng.integers(-4, 4)))
        trp = max(50, min(100, qs + rng.integers(-2, 2)))
        dispo = max(40, min(95, 80 - (dmt - dmt_cible_demo) * 0.1))
        lignes.append(
            {
                "TRANCHE": h.strftime("%H:%M"),
                "REÇUS": recus,
                "TRAITÉS": traites,
                "PRÉVISION": prevision,
                "TRP %": round(trp, 1),
                "QS %": round(qs, 1),
                "SL %": round(sl, 1),
                "DMC (S)": round(dmt * 0.9, 0),
                "ACW (S)": round(rng.uniform(10, 22), 0),
                "DMT (S)": round(dmt, 0),
                "CONNECTÉS": connectes,
                "EN TRAIT.": max(connectes - rng.integers(0, 2), 0),
                "DISPO %": round(dispo, 1),
                "ABAND. MOY": round(rng.uniform(2, 8), 1),
                "BESOIN": planning + rng.integers(0, 2),
                "PLANNING": planning,
                "ÉCART": connectes - planning,
            }
        )
    return pd.DataFrame(lignes)


def generer_demo_planning(n_agents: int = 28) -> pd.DataFrame:
    """Génère un planning fictif d'agents avec heures de prise de poste et
    pauses déjeuner étalées entre 11:30 et 14:00."""
    rng = np.random.default_rng(7)
    lignes = []
    debuts_possibles = ["07:30", "08:00", "08:30", "09:00", "10:00"]
    fins_possibles = ["16:30", "17:00", "17:30", "18:00", "19:00"]
    pauses_possibles = ["11:30", "12:00", "12:30", "13:00", "13:30"]
    for i in range(n_agents):
        hd = rng.choice(debuts_possibles)
        hf = rng.choice(fins_possibles)
        pd_ = rng.choice(pauses_possibles)
        h, m = map(int, pd_.split(":"))
        fin_pause = dt.time(h, m + 45) if m + 45 < 60 else dt.time(h + 1, m + 45 - 60)
        lignes.append(
            {
                "Login_Vocalcom": f"AG{i+1:03d}",
                "Code_RH": f"RH{1000+i}",
                "Nom_Prenom": f"Agent {i+1:02d}",
                "Heure_Debut": hd,
                "Heure_Fin": hf,
                "Pause_Debut": pd_,
                "Pause_Fin": fin_pause.strftime("%H:%M"),
            }
        )
    return pd.DataFrame(lignes)


# =============================================================================
# 4. BARRE LATÉRALE — CONFIGURATION DES OBJECTIFS & CHARGEMENT DES FICHIERS
# =============================================================================
with st.sidebar:
    st.markdown("## ⚙️ Configuration des objectifs")

    obj_sl = st.slider("Objectif SL — Service Level (%)", 50, 100, 85)
    seuil_reponse = st.number_input("Seuil de réponse SL (secondes)", min_value=5, value=20, step=5)
    obj_qs = st.slider("Objectif QS / TRP (%)", 50, 100, 95)
    dmt_cible = st.number_input("DMT Cible / Prévue (secondes)", min_value=30, value=280, step=10)
    acw_cible = st.number_input("ACW Cible (secondes)", min_value=0, value=15, step=1)
    capacite_rattrapage = st.slider(
        "Capacité max de rattrapage SL sur le reste de la journée (%)", 50, 100, 95
    ) / 100.0

    st.markdown("---")
    st.markdown("## 📁 Données de production")
    fichier_prod = st.file_uploader(
        "Fichier 1 — Métriques de production par tranche (Excel/CSV)",
        type=["xlsx", "xls", "csv"],
        key="fichier_prod",
    )
    st.download_button(
        "📄 Télécharger le modèle — Production",
        data=generer_modele_production(),
        file_name="modele_production_tranches.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_modele_prod",
        width='stretch',
    )

    fichier_planning = st.file_uploader(
        "Fichier 2 — Planning agents & pauses (Excel/CSV)",
        type=["xlsx", "xls", "csv"],
        key="fichier_planning",
    )
    st.download_button(
        "📄 Télécharger le modèle — Planning",
        data=generer_modele_planning(),
        file_name="modele_planning_agents.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_modele_planning",
        width='stretch',
    )

    mode_demo = fichier_prod is None or fichier_planning is None
    if mode_demo:
        st.info(
            "Aucun fichier chargé (ou fichier manquant) — l'application "
            "fonctionne actuellement sur un **jeu de données de démonstration** "
            "illustrant une dérive DMT en milieu de journée.",
            icon="ℹ️",
        )

    st.markdown("---")
    seuil_alerte_marge = st.slider(
        "Marge (points de SL) déclenchant l'état 'EN DANGER'", 1, 15, 4
    )
    seuil_derive_dmt_pct = st.slider("Seuil de dérive DMT significatif (%)", 5, 30, 10) / 100.0
    seuil_derive_flux_pct = st.slider("Seuil de surflux significatif (%)", 5, 30, 10) / 100.0
    seuil_sous_effectif = st.slider("Seuil sous-effectif (nb agents)", 1, 5, 2)


# =============================================================================
# 5. CHARGEMENT DES DONNÉES (fichiers réels ou démonstration)
# =============================================================================
if fichier_prod is not None:
    df_prod = charger_fichier(fichier_prod, MAP_COLONNES_PROD)
else:
    df_prod = generer_demo_production()
    df_prod.columns = [normaliser_nom_colonne(c) for c in df_prod.columns]
    df_prod = df_prod.rename(columns={k: v for k, v in MAP_COLONNES_PROD.items() if k in df_prod.columns})

if fichier_planning is not None:
    df_planning = charger_fichier(fichier_planning, MAP_COLONNES_PLANNING)
else:
    df_planning = generer_demo_planning()
    df_planning.columns = [normaliser_nom_colonne(c) for c in df_planning.columns]
    df_planning = df_planning.rename(
        columns={k: v for k, v in MAP_COLONNES_PLANNING.items() if k in df_planning.columns}
    )

# --- Nettoyage / typage -----------------------------------------------------
for col_num in ["recus", "traites", "prevision", "dmc", "acw", "dmt", "connectes",
                 "en_traitement", "aband_moy", "besoin", "planning", "ecart"]:
    if col_num in df_prod.columns:
        df_prod[col_num] = pd.to_numeric(df_prod[col_num], errors="coerce").fillna(0.0)

for col_pct in ["trp", "qs", "sl", "dispo"]:
    if col_pct in df_prod.columns:
        df_prod[col_pct] = normaliser_pourcentage(df_prod[col_pct])

df_prod["heure_tranche"] = df_prod["tranche"].apply(parse_heure)
df_prod = df_prod.dropna(subset=["heure_tranche"]).sort_values("heure_tranche").reset_index(drop=True)
# Ré-écrit la colonne 'tranche' en texte HH:MM lisible et JSON-sérialisable,
# quel que soit le type d'origine (str, datetime.time, Timestamp Excel...).
df_prod["tranche"] = df_prod["heure_tranche"].apply(lambda h: h.strftime("%H:%M"))

for col_h in ["heure_debut", "heure_fin", "pause_debut", "pause_fin"]:
    if col_h in df_planning.columns:
        df_planning[col_h] = df_planning[col_h].apply(parse_heure)


# =============================================================================
# 6. MOTEUR WFM — CUMULS, PROJECTION, POINT DE NON-RETOUR, DIAGNOSTIC
# =============================================================================
def compter_effectif_planning(df_planning: pd.DataFrame, debut: dt.time, fin: dt.time):
    """Pour une tranche [debut, fin), retourne (nb agents planifiés présents,
    nb agents en pause déjeuner sur ce créneau)."""
    n_planifies, n_pause = 0, 0
    for _, agent in df_planning.iterrows():
        hd, hf = agent.get("heure_debut"), agent.get("heure_fin")
        if hd is None or hf is None:
            continue
        if hd <= debut and hf >= fin:
            n_planifies += 1
            p_deb, p_fin = agent.get("pause_debut"), agent.get("pause_fin")
            if p_deb is not None and p_fin is not None:
                chevauchement = not (p_fin <= debut or p_deb >= fin)
                if chevauchement:
                    n_pause += 1
    return n_planifies, n_pause


def diagnostiquer_cause_racine(row: pd.Series, dmt_cible, seuil_derive_dmt_pct,
                                seuil_derive_flux_pct, seuil_sous_effectif):
    """Arbre de décision de diagnostic de cause racine pour une tranche donnée.
    Retourne (code_diagnostic, libelle, flags)."""
    ecart_dmt_pct = (row["dmt"] - dmt_cible) / dmt_cible if dmt_cible else 0.0
    effectif_dispo = max(row["planning"] - row["agents_en_pause"], 0)
    ecart_effectif = row["connectes"] - effectif_dispo
    ecart_flux_pct = (row["recus"] - row["prevision"]) / row["prevision"] if row["prevision"] else 0.0

    flag_dmt = ecart_dmt_pct > seuil_derive_dmt_pct
    flag_effectif = ecart_effectif < -seuil_sous_effectif
    flag_flux = ecart_flux_pct > seuil_derive_flux_pct

    causes = []
    if flag_dmt:
        causes.append(f"DMT +{ecart_dmt_pct*100:.0f}% vs cible")
    if flag_effectif:
        causes.append(f"Sous-effectif {ecart_effectif:.0f} agent(s) vs planifié")
    if flag_flux:
        causes.append(f"Surflux +{ecart_flux_pct*100:.0f}% vs prévision")

    if sum([flag_dmt, flag_effectif, flag_flux]) >= 2:
        code = "COMBINAISON"
        libelle = "Combinaison de facteurs (" + " + ".join(causes) + ")"
    elif flag_dmt:
        code = "DMT"
        libelle = f"Dérive opérationnelle sur le temps de traitement (DMT +{ecart_dmt_pct*100:.0f}%)"
    elif flag_effectif:
        code = "EFFECTIF"
        libelle = f"Sous-effectif / inadhérence ou retards plateau ({ecart_effectif:.0f} agent(s) manquant(s))"
    elif flag_flux:
        code = "FLUX"
        libelle = f"Surflux volumétrique imprévu — Out of Forecast (+{ecart_flux_pct*100:.0f}%)"
    else:
        code = "NOMINAL"
        libelle = "Situation nominale — aucune dérive significative détectée"

    flags = {"dmt": flag_dmt, "effectif": flag_effectif, "flux": flag_flux}
    return code, libelle, flags


ACTIONS_PRESCRIPTIVES = {
    "DMT": [
        "Activer les consignes d'écourtage DMT sur les motifs longs",
        "Rappel flash aux superviseurs sur la maîtrise du temps de traitement",
        "Basculer les appels complexes vers les agents seniors / experts",
    ],
    "EFFECTIF": [
        "Décaler ou raccourcir les pauses de 15 min sur les 2 prochaines tranches",
        "Rappeler les agents en formation / coaching disponibles",
        "Solliciter du renfort en heures supplémentaires ou back-up",
    ],
    "FLUX": [
        "Activer un message d'attente / SVI de débordement",
        "Basculer une partie du flux vers un canal différé (email, callback)",
        "Solliciter un plateau de débordement si accord multi-site",
    ],
    "COMBINAISON": [
        "Déclencher la cellule de crise pilotage (WFM + Superviseurs)",
        "Prioriser : stabiliser l'effectif AVANT d'agir sur la DMT",
        "Réévaluer l'objectif SL horaire restant en concertation client",
    ],
    "NOMINAL": [
        "Maintenir la vigilance sur les 2 prochaines tranches",
        "Aucune action corrective nécessaire à ce stade",
    ],
}


def moteur_wfm(df_prod: pd.DataFrame, df_planning: pd.DataFrame, obj_sl_pct: float,
               capacite_rattrapage: float, dmt_cible, seuil_alerte_marge,
               seuil_derive_dmt_pct, seuil_derive_flux_pct, seuil_sous_effectif):
    """Cœur du moteur WFM : calcule pour chaque tranche les cumuls, la
    projection de fin de journée, le point de non-retour et le diagnostic."""
    df = df_prod.copy().reset_index(drop=True)
    obj_sl_frac = obj_sl_pct / 100.0

    # --- Effectif en pause par tranche (à partir du fichier planning) -------
    agents_en_pause = []
    for _, row in df.iterrows():
        debut = row["heure_tranche"]
        fin_dt = (dt.datetime.combine(dt.date.today(), debut) + dt.timedelta(minutes=30)).time()
        _, n_pause = compter_effectif_planning(df_planning, debut, fin_dt)
        agents_en_pause.append(n_pause)
    df["agents_en_pause"] = agents_en_pause

    # --- Cumuls réalisés ------------------------------------------------------
    df["cum_recus"] = df["recus"].cumsum()
    df["conformes_tranche"] = df["recus"] * (df["sl"] / 100.0)
    df["cum_conformes"] = df["conformes_tranche"].cumsum()
    df["sl_cumule"] = np.where(df["cum_recus"] > 0, df["cum_conformes"] / df["cum_recus"] * 100.0, 0.0)

    # --- Prévision restante (tranches futures, strictement après la tranche) -
    total_prevision = df["prevision"].sum()
    df["prevision_cumulee"] = df["prevision"].cumsum()
    df["prevision_future"] = total_prevision - df["prevision_cumulee"]

    # --- SL Max Projeté en fin de journée & Point de Non-Retour --------------
    denominateur = df["cum_recus"] + df["prevision_future"]
    numerateur = df["cum_conformes"] + df["prevision_future"] * capacite_rattrapage
    df["sl_max_projete"] = np.where(denominateur > 0, numerateur / denominateur * 100.0, df["sl_cumule"])

    df["point_non_retour"] = df["sl_max_projete"] < obj_sl_pct

    # --- Statut par tranche (vert / orange / rouge) ---------------------------
    def determiner_statut(sl_projete):
        if sl_projete < obj_sl_pct:
            return "POINT DE NON RETOUR ATTEINT"
        elif sl_projete < obj_sl_pct + seuil_alerte_marge:
            return "EN DANGER"
        else:
            return "SOUS CONTROLE"

    df["statut"] = df["sl_max_projete"].apply(determiner_statut)

    # --- Diagnostic de cause racine par tranche -------------------------------
    codes, libelles = [], []
    for _, row in df.iterrows():
        code, libelle, _ = diagnostiquer_cause_racine(
            row, dmt_cible, seuil_derive_dmt_pct, seuil_derive_flux_pct, seuil_sous_effectif
        )
        codes.append(code)
        libelles.append(libelle)
    df["diagnostic_code"] = codes
    df["diagnostic_libelle"] = libelles

    # --- Heure exacte de bascule (première tranche en point de non-retour) ---
    heure_bascule = None
    tranches_bascule = df.loc[df["point_non_retour"], "tranche"]
    if not tranches_bascule.empty:
        heure_bascule = tranches_bascule.iloc[0]

    synthese = {
        "heure_bascule": heure_bascule,
        "statut_actuel": df["statut"].iloc[-1],
        "sl_cumule_actuel": df["sl_cumule"].iloc[-1],
        "sl_max_projete_actuel": df["sl_max_projete"].iloc[-1],
        "diagnostic_code_actuel": df["diagnostic_code"].iloc[-1],
        "diagnostic_libelle_actuel": df["diagnostic_libelle"].iloc[-1],
        "derniere_tranche": df["tranche"].iloc[-1],
    }
    return df, synthese


df_result, synthese = moteur_wfm(
    df_prod, df_planning, obj_sl, capacite_rattrapage, dmt_cible, seuil_alerte_marge,
    seuil_derive_dmt_pct, seuil_derive_flux_pct, seuil_sous_effectif,
)

actions_actuelles = ACTIONS_PRESCRIPTIVES.get(synthese["diagnostic_code_actuel"], [])


# =============================================================================
# 7. COMPOSANTS UI — CONTROL ROOM (HTML / CSS / JS injectés)
# =============================================================================
def render_banniere_alerte(statut: str, heure_bascule, sl_actuel: float, sl_projete: float, obj_sl: float):
    """Bandeau supérieur dynamique — vert / orange / rouge clignotant néon."""
    if statut == "POINT DE NON RETOUR ATTEINT":
        couleur, icone, anim = "#ff2d55", "🚨", "blink-neon 1.1s infinite"
        message = f"POINT DE NON-RETOUR ATTEINT À {heure_bascule} — OBJECTIF SL {obj_sl:.0f}% MATHÉMATIQUEMENT INATTEIGNABLE"
    elif statut == "EN DANGER":
        couleur, icone, anim = "#ffae00", "⚠️", "pulse-orange 1.6s infinite"
        message = "SL EN DANGER — MARGE RÉSIDUELLE FAIBLE AVANT LE POINT DE NON-RETOUR"
    else:
        couleur, icone, anim = "#00ff88", "✅", "none"
        message = "SITUATION SOUS CONTRÔLE — OBJECTIF SL ATTEIGNABLE"

    tpl = string.Template(
        """
        <style>
        @keyframes blink-neon {
            0%, 100% { opacity:1; box-shadow:0 0 20px 6px $couleur; }
            50% { opacity:0.55; box-shadow:0 0 55px 16px $couleur; }
        }
        @keyframes pulse-orange {
            0%, 100% { box-shadow:0 0 10px 2px $couleur; }
            50% { box-shadow:0 0 28px 8px $couleur; }
        }
        .banniere {
            background: linear-gradient(90deg, #0a0e17 0%, ${couleur}22 100%);
            border: 2px solid $couleur;
            border-radius: 14px;
            padding: 18px 26px;
            display: flex; align-items: center; justify-content: space-between;
            font-family: 'Segoe UI', sans-serif; color: #f2f6ff;
            animation: $anim;
        }
        .banniere-txt { font-size: 19px; font-weight: 800; letter-spacing: 0.4px; }
        .banniere-sub { font-size: 13px; opacity: 0.85; margin-top: 6px; color:#b9c6de;}
        .banniere-metric { font-size: 30px; font-weight: 900; color: $couleur; text-align:right; }
        .banniere-metric-label { font-size: 11px; color:#8fa3c7; text-align:right; letter-spacing:1px;}
        </style>
        <div class="banniere">
          <div>
            <div class="banniere-txt">$icone $message</div>
            <div class="banniere-sub">SL cumulé réalisé : $sl_actuel% &nbsp;|&nbsp; SL max projeté fin de journée : $sl_projete% &nbsp;|&nbsp; Objectif : $obj_sl%</div>
          </div>
          <div>
            <div class="banniere-metric">$sl_projete%</div>
            <div class="banniere-metric-label">SL PROJETÉ</div>
          </div>
        </div>
        """
    )
    html = tpl.substitute(
        couleur=couleur, icone=icone, anim=anim, message=message,
        sl_actuel=f"{sl_actuel:.1f}", sl_projete=f"{sl_projete:.1f}", obj_sl=f"{obj_sl:.0f}",
    )
    components.html(html, height=115)


def render_jauge_sl(sl_actuel: float, sl_projete: float, obj_sl: float):
    """Jauge à aiguille (Chart.js + plugin custom) affichant le SL max projeté,
    avec repères SL actuel / objectif.
    Le canvas est enveloppé dans un conteneur à hauteur fixe et
    maintainAspectRatio est désactivé pour un rendu pleinement responsive
    (plus besoin de zoomer pour voir le graphique correctement)."""
    zone_rouge = max(obj_sl - 15, 0)
    zone_orange = 15
    zone_verte = max(100 - obj_sl, 0)

    tpl = string.Template(
        """
        <div style="background:#0a0e17; border:1px solid #1c2536; border-radius:16px; padding:14px; box-sizing:border-box;">
          <div style="position:relative; width:100%; height:250px;">
            <canvas id="gaugeChart"></canvas>
          </div>
          <div style="display:flex; justify-content:space-around; margin-top:6px; font-family:'Segoe UI',sans-serif;">
            <div style="text-align:center;"><div style="color:#8fa3c7;font-size:11px;">SL ACTUEL CUMULÉ</div><div style="color:#00e5ff;font-size:20px;font-weight:800;">$sl_actuel%</div></div>
            <div style="text-align:center;"><div style="color:#8fa3c7;font-size:11px;">SL MAX PROJETÉ</div><div style="color:#ffffff;font-size:20px;font-weight:800;">$sl_projete%</div></div>
            <div style="text-align:center;"><div style="color:#8fa3c7;font-size:11px;">OBJECTIF</div><div style="color:#00ff88;font-size:20px;font-weight:800;">$obj_sl%</div></div>
          </div>
        </div>
        <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
        <script>
        const ctx = document.getElementById('gaugeChart').getContext('2d');
        const valeur = $sl_projete_raw;

        const aiguillePlugin = {
          id: 'aiguillePlugin',
          afterDatasetsDraw(chart) {
            const meta = chart.getDatasetMeta(0);
            const centre = meta.data[0];
            if (!centre) return;
            const cx = centre.x, cy = centre.y;
            const rayon = (centre.outerRadius) * 0.92;
            const angle = Math.PI * (1 - (valeur / 100));
            const x2 = cx + rayon * Math.cos(angle);
            const y2 = cy - rayon * Math.sin(angle);
            const c = chart.ctx;
            c.save();
            c.strokeStyle = '#ffffff';
            c.lineWidth = 4;
            c.beginPath();
            c.moveTo(cx, cy);
            c.lineTo(x2, y2);
            c.stroke();
            c.beginPath();
            c.fillStyle = '#ffffff';
            c.arc(cx, cy, 7, 0, 2 * Math.PI);
            c.fill();
            c.restore();
          }
        };

        new Chart(ctx, {
          type: 'doughnut',
          data: {
            labels: ['Zone rouge', 'Zone orange', 'Zone verte'],
            datasets: [{
              data: [$zone_rouge, $zone_orange, $zone_verte],
              backgroundColor: ['#ff2d55', '#ffae00', '#00ff88'],
              borderWidth: 0,
            }]
          },
          options: {
            responsive: true,
            maintainAspectRatio: false,
            circumference: 180,
            rotation: -90,
            cutout: '68%',
            plugins: { legend: { display: false }, tooltip: { enabled: false } },
            animation: { animateRotate: true, duration: 900 },
          },
          plugins: [aiguillePlugin],
        });

        // Redessine proprement la jauge si l'iframe est redimensionnée
        // (ex: repli/dépli de la sidebar Streamlit).
        window.addEventListener('resize', () => {
          Chart.getChart('gaugeChart')?.resize();
        });
        </script>
        """
    )
    html = tpl.substitute(
        sl_actuel=f"{sl_actuel:.1f}", sl_projete=f"{sl_projete:.1f}", obj_sl=f"{obj_sl:.0f}",
        sl_projete_raw=f"{sl_projete:.2f}", zone_rouge=f"{zone_rouge:.1f}",
        zone_orange=f"{zone_orange:.1f}", zone_verte=f"{zone_verte:.1f}",
    )
    components.html(html, height=380)


def render_courbe_projection(df: pd.DataFrame, obj_sl: float):
    """Courbe interactive Chart.js : SL cumulé réalisé vs SL max projeté vs
    ligne d'objectif, tranche par tranche.
    Même correction que la jauge : conteneur à hauteur fixe +
    maintainAspectRatio désactivé pour un rendu net sans avoir à zoomer."""
    labels = [str(v) for v in df["tranche"].tolist()]
    sl_cumule = [float(v) for v in df["sl_cumule"].round(1).tolist()]
    sl_projete = [float(v) for v in df["sl_max_projete"].round(1).tolist()]
    objectif = [round(float(obj_sl), 1)] * len(df)
    bascule_idx = df.index[df["point_non_retour"]].tolist()
    bascule_x = labels[bascule_idx[0]] if bascule_idx else None

    tpl = string.Template(
        """
        <div style="background:#0a0e17; border:1px solid #1c2536; border-radius:16px; padding:16px; box-sizing:border-box;">
          <div style="position:relative; width:100%; height:290px;">
            <canvas id="courbeChart"></canvas>
          </div>
        </div>
        <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
        <script>
        const ctx2 = document.getElementById('courbeChart').getContext('2d');
        new Chart(ctx2, {
          type: 'line',
          data: {
            labels: $labels,
            datasets: [
              {
                label: 'SL cumulé réalisé (%)',
                data: $sl_cumule,
                borderColor: '#00e5ff',
                backgroundColor: 'rgba(0,229,255,0.08)',
                borderWidth: 2, tension: 0.25, fill: true, pointRadius: 2,
              },
              {
                label: 'SL max projeté fin de journée (%)',
                data: $sl_projete,
                borderColor: '#ffae00',
                borderWidth: 2, borderDash: [6, 4], tension: 0.25, pointRadius: 2, fill: false,
              },
              {
                label: 'Objectif SL (%)',
                data: $objectif,
                borderColor: '#00ff88',
                borderWidth: 2, borderDash: [2, 2], pointRadius: 0, fill: false,
              },
            ]
          },
          options: {
            responsive: true,
            maintainAspectRatio: false,
            interaction: { mode: 'index', intersect: false },
            plugins: {
              legend: { labels: { color: '#c7d3e8', font: { size: 11 } } },
              tooltip: { mode: 'index', intersect: false },
            },
            scales: {
              x: { ticks: { color: '#8fa3c7' }, grid: { color: '#141b2b' } },
              y: { ticks: { color: '#8fa3c7' }, grid: { color: '#141b2b' }, suggestedMin: 0, suggestedMax: 100 },
            },
          }
        });

        window.addEventListener('resize', () => {
          Chart.getChart('courbeChart')?.resize();
        });
        </script>
        """
    )
    html = tpl.substitute(
        labels=json.dumps(labels), sl_cumule=json.dumps(sl_cumule),
        sl_projete=json.dumps(sl_projete), objectif=json.dumps(objectif),
    )
    components.html(html, height=340)
    if bascule_x:
        st.caption(f"⛔ Bascule projetée au point de non-retour visible à la tranche **{bascule_x}**.")


def render_vigie_ia(diagnostic_libelle: str, actions: list, heure_analyse: str, statut: str):
    """Panneau 'Vigie IA Assistant' avec effet machine à écrire (JavaScript).

    Corrections apportées :
      - le texte affiché ne contient plus de mention figée du type
        « RECOMMANDATIONS EN COURS DE GÉNÉRATION... » qui ne disparaissait
        jamais une fois l'animation terminée (donnait l'impression que la
        Vigie était bloquée / ne fonctionnait pas) ;
      - les actions prescriptives apparaissent désormais en cascade
        (délai croissant) une fois le texte d'analyse entièrement écrit,
        au lieu de toutes s'afficher en même temps.
    """
    couleur = {"POINT DE NON RETOUR ATTEINT": "#ff2d55", "EN DANGER": "#ffae00"}.get(statut, "#00ff88")
    texte_analyse = (
        f"[ANALYSE VIGIE — TRANCHE {heure_analyse}]\n"
        f"> STATUT : {statut}\n"
        f"> DIAGNOSTIC : {diagnostic_libelle}"
    )
    actions_html = "".join(f'<div class="vigie-action">▸ {a}</div>' for a in actions)

    tpl = string.Template(
        """
        <style>
        .vigie-panel {
            background: #060a12; border: 1px solid $couleur; border-radius: 14px;
            padding: 18px 22px; font-family: 'Consolas', 'Courier New', monospace;
            box-shadow: 0 0 18px 0 ${couleur}33;
        }
        .vigie-header { color: $couleur; font-size: 14px; font-weight: 800; letter-spacing: 1px; margin-bottom: 10px; }
        .vigie-text { color: #c9f0ff; font-size: 13.5px; line-height: 1.7; white-space: pre-wrap; min-height: 70px; }
        .vigie-actions { margin-top: 14px; border-top: 1px dashed #223049; padding-top: 10px; }
        .vigie-action {
            color: #e6f1ff; font-size: 13px; padding: 3px 0;
            opacity: 0; animation: apparition 0.5s forwards;
            animation-play-state: paused;
        }
        .vigie-action.pret { animation-play-state: running; }
        @keyframes apparition { to { opacity: 1; } }
        .curseur { display:inline-block; width:8px; background:$couleur; animation: clignote 0.8s infinite; }
        @keyframes clignote { 0%,100% {opacity:1;} 50% {opacity:0;} }
        </style>
        <div class="vigie-panel">
          <div class="vigie-header">🤖 VIGIE IA ASSISTANT — ANALYSE PRÉDICTIVE & CAUSE RACINE</div>
          <div id="vigie-text" class="vigie-text"></div>
          <div class="vigie-actions">$actions_html</div>
        </div>
        <script>
        const texte = $texte_json;
        let i = 0;
        const el = document.getElementById('vigie-text');

        function ecrire() {
            if (i < texte.length) {
                el.innerHTML = texte.substring(0, i + 1).replace(/\\n/g, '<br>') + '<span class="curseur">&nbsp;</span>';
                i++;
                setTimeout(ecrire, 14);
            } else {
                // Texte final, sans curseur clignotant résiduel
                el.innerHTML = texte.replace(/\\n/g, '<br>');
                // Révèle les actions en cascade, une fois l'analyse terminée
                const actions = document.querySelectorAll('.vigie-action');
                actions.forEach((a, idx) => {
                    a.style.animationDelay = (idx * 0.18) + 's';
                    a.classList.add('pret');
                });
            }
        }
        ecrire();
        </script>
        """
    )
    html = tpl.substitute(
        couleur=couleur, actions_html=actions_html, texte_json=json.dumps(texte_analyse),
    )
    components.html(html, height=210 + max(len(actions), 1) * 26)


def construire_dataframe_affichage(df: pd.DataFrame) -> pd.DataFrame:
    """Construit le dataframe final avec libellés français pour l'affichage."""
    df_aff = pd.DataFrame({
        "Tranche": df["tranche"],
        "Reçus": df["recus"].astype(int),
        "Traités": df["traites"].astype(int),
        "Prévision": df["prevision"].astype(int),
        "TRP %": df["trp"].round(1),
        "QS %": df["qs"].round(1),
        "SL %": df["sl"].round(1),
        "DMT (S)": df["dmt"].round(0).astype(int),
        "ACW (S)": df["acw"].round(0).astype(int),
        "Connectés": df["connectes"].astype(int),
        "Planning": df["planning"].astype(int),
        "En pause": df["agents_en_pause"].astype(int),
        "Écart Effectif": (df["connectes"] - (df["planning"] - df["agents_en_pause"])).astype(int),
        "SL Cumulé %": df["sl_cumule"].round(1),
        "SL Max Projeté %": df["sl_max_projete"].round(1),
        "Statut": df["statut"],
        "Diagnostic": df["diagnostic_libelle"],
    })
    return df_aff


def styliser_dataframe(df_aff: pd.DataFrame, obj_sl: float, obj_qs: float, dmt_cible: float,
                        seuil_alerte_marge: float):
    """Applique une mise en forme conditionnelle (heatmap vert/jaune/rouge)."""

    def couleur_pct(val, objectif):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return ""
        if v >= objectif:
            return "background-color:#0a3d2c;color:#4dffb0;"
        elif v >= objectif - 10:
            return "background-color:#4a3a05;color:#ffd166;"
        else:
            return "background-color:#4a0a1c;color:#ff6b8b;"

    def couleur_dmt(val):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return ""
        ecart = (v - dmt_cible) / dmt_cible if dmt_cible else 0
        if ecart <= 0.05:
            return "background-color:#0a3d2c;color:#4dffb0;"
        elif ecart <= seuil_derive_dmt_pct:
            return "background-color:#4a3a05;color:#ffd166;"
        else:
            return "background-color:#4a0a1c;color:#ff6b8b;"

    def couleur_effectif(val):
        try:
            v = float(val)
        except (TypeError, ValueError):
            return ""
        if v >= 0:
            return "background-color:#0a3d2c;color:#4dffb0;"
        elif v >= -seuil_sous_effectif:
            return "background-color:#4a3a05;color:#ffd166;"
        else:
            return "background-color:#4a0a1c;color:#ff6b8b;"

    def couleur_statut(val):
        if val == "POINT DE NON RETOUR ATTEINT":
            return "background-color:#4a0a1c;color:#ff6b8b;font-weight:700;"
        elif val == "EN DANGER":
            return "background-color:#4a3a05;color:#ffd166;font-weight:700;"
        else:
            return "background-color:#0a3d2c;color:#4dffb0;font-weight:700;"

    styler = (
        df_aff.style
        .map(lambda v: couleur_pct(v, obj_sl), subset=["SL %", "SL Cumulé %", "SL Max Projeté %"])
        .map(lambda v: couleur_pct(v, obj_qs), subset=["QS %", "TRP %"])
        .map(couleur_dmt, subset=["DMT (S)"])
        .map(couleur_effectif, subset=["Écart Effectif"])
        .map(couleur_statut, subset=["Statut"])
        .set_properties(**{"font-family": "Segoe UI, sans-serif", "font-size": "12.5px"})
    )
    return styler


# =============================================================================
# 8. MISE EN PAGE PRINCIPALE
# =============================================================================
st.markdown("# 📡 CONTROL ROOM — Pilotage Production Temps Réel & Prédictif")
st.caption(
    "Détection du point de non-retour SL · Diagnostic automatique de cause racine · "
    "Actions prescriptives Vigie"
)

render_banniere_alerte(
    synthese["statut_actuel"], synthese["heure_bascule"],
    synthese["sl_cumule_actuel"], synthese["sl_max_projete_actuel"], obj_sl,
)

st.markdown('<div class="bloc-titre">Indicateurs clés — dernière tranche analysée</div>', unsafe_allow_html=True)
k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("SL cumulé réalisé", f"{synthese['sl_cumule_actuel']:.1f}%")
k2.metric("SL max projeté", f"{synthese['sl_max_projete_actuel']:.1f}%",
          delta=f"{synthese['sl_max_projete_actuel'] - obj_sl:.1f} pts vs objectif")
k3.metric("Objectif SL", f"{obj_sl:.0f}%")
heure_b = str(synthese["heure_bascule"]) if synthese.get("heure_bascule") and pd.notna(synthese["heure_bascule"]) else "Non atteinte"
k4.metric("Heure de bascule", heure_b)
k5.metric("Appels reçus (cumul)", f"{int(df_result['cum_recus'].iloc[-1])}")

st.markdown("---")

col_g, col_c = st.columns([1, 2])
with col_g:
    st.markdown('<div class="bloc-titre">Jauge de survie SL</div>', unsafe_allow_html=True)
    render_jauge_sl(synthese["sl_cumule_actuel"], synthese["sl_max_projete_actuel"], obj_sl)
with col_c:
    st.markdown('<div class="bloc-titre">Projection SL — tranche par tranche</div>', unsafe_allow_html=True)
    render_courbe_projection(df_result, obj_sl)

st.markdown('<div class="bloc-titre">Vigie IA — diagnostic & actions prescriptives</div>', unsafe_allow_html=True)
render_vigie_ia(
    synthese["diagnostic_libelle_actuel"], actions_actuelles,
    synthese["derniere_tranche"], synthese["statut_actuel"],
)

st.markdown("---")
st.markdown('<div class="bloc-titre">Tableau de bord détaillé — toutes tranches</div>', unsafe_allow_html=True)

df_affichage = construire_dataframe_affichage(df_result)
styler = styliser_dataframe(df_affichage, obj_sl, obj_qs, dmt_cible, seuil_alerte_marge)
st.dataframe(styler, width='stretch', height=480)

# --- Export des résultats ----------------------------------------------------
csv_export = df_affichage.to_csv(index=False, sep=";").encode("utf-8-sig")
st.download_button(
    "⬇️ Exporter le tableau de bord (CSV)",
    data=csv_export,
    file_name=f"control_room_wfm_{dt.date.today().isoformat()}.csv",
    mime="text/csv",
)

with st.expander("ℹ️ Méthodologie du moteur prédictif"):
    st.markdown(
        """
        **Point de non-retour** : à chaque tranche, le SL maximum théoriquement
        atteignable en fin de journée est recalculé selon :

        `SL_Max_Projeté = (Conformes_cumulés + Prévision_future × Capacité_rattrapage) / (Reçus_cumulés + Prévision_future)`

        Dès que cette valeur passe sous l'objectif SL, la tranche est marquée
        comme point de non-retour : même en traitant 100 % du flux restant
        dans les temps, l'objectif de la journée ne peut plus être atteint.

        **Diagnostic de cause racine** : à chaque tranche, l'écart de DMT,
        l'écart d'effectif disponible (planning − agents en pause vs agents
        connectés) et l'écart de flux (reçus vs prévision) sont comparés à des
        seuils paramétrables. Le diagnostic combine ces facteurs selon l'arbre
        de décision défini dans la barre latérale.
        """
    )
